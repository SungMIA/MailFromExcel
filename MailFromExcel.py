# Sungdong Kim
# sungkimmia@gmail.com
# 08/12/2018
# MailFromExcel.py 
# Automatically Send Customized Email Messages with Addresses from an Excel Spreadsheet

import smtplib
import openpyxl
import sys

wb = openpyxl.load_workbook('Spreadsheet.xlsx')
#Replace with file name of spreadsheet

sheet = wb['Sheet1']

recipients = {}
for r in range(2, sheet.max_row + 1):
	name = sheet.cell(row=r, column=1).value
	email = sheet.cell(row=r, column=4).value
	recipients[name] = email
#Change in relation to the location of the data in the excel file	



server = smtplib.SMTP('smtp.gmail.com', 587)
#The sender's email type. Replace with:
#Gmail: smtp.gmail.com
#Outlook/Hotmail: smtp-mail.outlook.com
#Yahoo: smtp.mail.yahoo.com

server.ehlo()
server.starttls()
server.login('username@gmail.com', 'password;')
#Sender Email Login Information
# REMEMBER : Delete your password from this program after using
# Protects privacy

for name, email in recipients.items():
	body = "From: %s\nTo: %s\nSubject: 'SUBJECT OF MAIL'\n\n'BODY'\n\n'CLOSING REMARKS'\n\n\n'SIGNATURE\n\n' % ('username@gmail.com', email, name)
	email = [email] + ['username@gmail.com']
	#BCC. In order to change BCC address, edit line 40
	server.sendmail('username@gmail.com', email, body)
server.quit()
